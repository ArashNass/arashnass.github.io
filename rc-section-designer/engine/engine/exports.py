"""
engine/exports.py
PDF and Excel report generation.
Called by routes/api.py — keep all reportlab/openpyxl logic here.
"""
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors as C
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── helpers ──────────────────────────────────────────────────────────
BLUE  = C.HexColor("#1a6cf5")
LBLUE = C.HexColor("#eaf1fe")
LGRAY = C.HexColor("#f0f3fa")
DGRAY = C.HexColor("#18243d")
GREEN = C.HexColor("#16a34a")
RED   = C.HexColor("#dc2626")
BORD  = C.HexColor("#dde3ef")

def _table(rows, widths, pass_col=None, pass_rows=None):
    sty = TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), BLUE),
        ("TEXTCOLOR",     (0,0), (-1,0), C.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [C.white, LGRAY]),
        ("GRID",          (0,0), (-1,-1), 0.4, BORD),
        ("ALIGN",         (1,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ])
    if pass_col is not None and pass_rows:
        for r, ok in pass_rows:
            sty.add("TEXTCOLOR", (pass_col,r), (pass_col,r), GREEN if ok else RED)
            sty.add("FONTNAME",  (pass_col,r), (pass_col,r), "Helvetica-Bold")
    t = Table(rows, colWidths=widths)
    t.setStyle(sty)
    return t


def _param_rows(params):
    shape = params.get("shape", "rect")
    rows = [["Design Code", params.get("code","—"), "—"],
            ["Section shape", {"rect":"Rectangular","circ":"Circular","tee":"T-section"}.get(shape, shape), "—"]]
    if shape == "circ":
        rows.append(["Diameter D", params.get("D","—"), "mm"])
    elif shape == "tee":
        rows += [["Flange width bf", params.get("bf","—"), "mm"],
                 ["Flange depth hf", params.get("hf","—"), "mm"],
                 ["Web width bw",    params.get("bw","—"), "mm"],
                 ["Depth h",         params.get("h","—"),  "mm"]]
    else:
        rows += [["Width b", params.get("b","—"), "mm"],
                 ["Depth h", params.get("h","—"), "mm"]]
    if params.get("code") == "EC2":
        rows += [["Partial factor γc",   params.get("gamma_c","1.50"), "—"],
                 ["Partial factor γs",   params.get("gamma_s","1.15"), "—"],
                 ["Coefficient αcc",     params.get("acc","0.85"),     "—"]]
    rows += [["Cover to bar centre d'", params.get("dp","—"),   "mm"],
             ["Concrete strength f'c",  params.get("fc","—"),   "MPa"],
             ["Steel yield strength fy",params.get("fy","—"),   "MPa"],
             ["Reinforcement ratio ρ",  params.get("rho","—"),  "%"],
             ["Bar diameter",           params.get("dbar","—"), "mm"],
             ["Design axial force N*",  params.get("Nd","—"),   "kN"],
             ["Design moment M*",       params.get("Md","—"),   "kNm"],
             ["Design shear V*",        params.get("Vd","—"),   "kN"],
             ["SLS moment M*sls",       params.get("Msls","—"), "kNm"]]
    sl = params.get("slender") or {}
    if sl.get("enabled"):
        rows += [["Unbraced length lu",  sl.get("lu","—"), "mm"],
                 ["Effective length k",  sl.get("k","—"),  "—"],
                 ["Frame type", "Sway" if sl.get("sway") else "Braced", "—"]]
    return rows


# ── PDF ───────────────────────────────────────────────────────────────
def generate_pdf(params, curve, checks, mc_data):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    S_title = ParagraphStyle("T", parent=styles["Normal"],
        fontSize=22, textColor=BLUE, fontName="Helvetica-Bold", spaceAfter=4, leading=26)
    S_sub   = ParagraphStyle("S", parent=styles["Normal"],
        fontSize=11, textColor=C.HexColor("#546080"), spaceAfter=2)
    S_h2    = ParagraphStyle("H", parent=styles["Normal"],
        fontSize=13, textColor=DGRAY, fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    S_note  = ParagraphStyle("N", parent=styles["Normal"],
        fontSize=9,  textColor=C.HexColor("#546080"), leading=13)

    story = []
    story.append(Paragraph("SectionForge", S_title))
    story.append(Paragraph("Reinforced Concrete Section Design Report", S_sub))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}", S_sub))
    story.append(Paragraph("© 2026 Arash Nassirpour", S_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=12))

    # Input parameters
    story.append(Paragraph("Section Parameters", S_h2))
    story.append(_table([["Parameter", "Value", "Unit"]] + _param_rows(params),
                        [8*cm, 5*cm, 3.5*cm]))

    # M-N results
    story.append(Paragraph("M – N Interaction Results", S_h2))
    uls  = checks.get("uls",{});  sh  = checks.get("shear",{})
    sls  = checks.get("sls",{});  rb  = checks.get("rebar",{})
    bal  = curve.get("bal",{})
    story.append(_table([
        ["Result",                "Value",                              "Status"],
        ["φPn,max (compression)", str(curve.get("Pmax","—"))+" kN",    "—"],
        ["Balanced axial force",  str(bal.get("y","—"))+" kN",         "—"],
        ["Balanced moment",       str(bal.get("x","—"))+" kNm",        "—"],
        ["Pure bending φMn",      str(curve.get("pure_M","—"))+" kNm", "—"],
        ["ULS utilisation",       str(uls.get("util","—"))+"%",        "PASS" if uls.get("pass") else "FAIL"],
        ["Neutral axis depth",    str(uls.get("c_mm","—"))+" mm",      "—"],
        ["φ factor",              str(uls.get("phi","—")),              "—"],
    ], [8*cm, 5*cm, 3.5*cm], pass_col=2, pass_rows=[(5, uls.get("pass",False))]))

    # Design checks
    story.append(Paragraph("Design Checks Summary", S_h2))
    story.append(_table([
        ["Check",           "Value",                             "Limit",                          "Status"],
        ["ULS flexure",     str(uls.get("util","—"))+"%",       "≤ 100%",                         "PASS" if uls.get("pass") else "FAIL"],
        ["Shear φVc",       str(sh.get("phiVc_kN","—"))+" kN", "V* = "+str(sh.get("Vd_kN","—"))+" kN", "PASS" if sh.get("pass") else "FAIL"],
        ["Crack width wk",  str(sls.get("wk_mm","—"))+" mm",   "≤ 0.3 mm",                       "PASS" if sls.get("pass") else "FAIL"],
        ["Steel area As",   str(rb.get("As_mm2","—"))+" mm²",  "min "+str(rb.get("As_min","—")), "PASS" if rb.get("pass_min") else "FAIL"],
    ], [6.5*cm, 4*cm, 4*cm, 2*cm], pass_col=3,
       pass_rows=[(i+1, row[3]=="PASS") for i,row in enumerate([
           ["","","","PASS" if uls.get("pass") else "FAIL"],
           ["","","","PASS" if sh.get("pass") else "FAIL"],
           ["","","","PASS" if sls.get("pass") else "FAIL"],
           ["","","","PASS" if rb.get("pass_min") else "FAIL"],
       ])]))

    # Slenderness
    sl = checks.get("slender") or {}
    if sl.get("enabled"):
        story.append(Paragraph("Slenderness (Second-Order Effects)", S_h2))
        rows = [["Parameter", "Value"],
                ["Slenderness ratio λ",  str(sl.get("lam","—"))],
                ["Slenderness limit",    str(sl.get("lam_lim","—"))]]
        if sl.get("magnified"):
            rows.append(["Magnified moment δM*", str(sl.get("Mc_kNm","—"))+" kNm"])
            if sl.get("delta") is not None: rows.append(["Magnifier δ", str(sl.get("delta"))])
            if sl.get("Pc_kN") is not None: rows.append(["Critical load Pc", str(sl.get("Pc_kN"))+" kN"])
            if sl.get("e2_mm") is not None: rows.append(["2nd-order ecc. e2", str(sl.get("e2_mm"))+" mm"])
        else:
            rows.append(["Result", "Short column — no magnification"])
        rows.append(["Method", sl.get("note","—")])
        story.append(_table(rows, [9*cm, 7.5*cm]))

    # Stirrups
    st = (checks.get("shear") or {}).get("stirrups")
    if st:
        story.append(Paragraph("Shear Link Design", S_h2))
        pick = st.get("pick")
        story.append(_table([
            ["Parameter", "Value"],
            ["Required Asw/s", str(st.get("Asw_s_mm2_per_m","—"))+" mm²/m"],
            ["Maximum spacing", str(st.get("s_max_mm","—"))+" mm"],
            ["Suggested links", f"{pick['legs']}-T{pick['dia']} @ {pick['s_mm']} mm" if pick else "see Asw/s"],
            ["Method", st.get("code","—")],
        ], [9*cm, 7.5*cm]))

    # M-phi
    if mc_data and mc_data.get("yield"):
        story.append(Paragraph("Moment – Curvature Summary", S_h2))
        y  = mc_data["yield"];  u = mc_data["ultimate"]
        mu = round(u["phi"]/y["phi"], 2) if y["phi"] > 0 else "N/A"
        story.append(_table([
            ["Parameter",          "Value"],
            ["Yield moment My",    str(y.get("M","—"))+" kNm"],
            ["Yield curvature φy", str(y.get("phi","—"))+" ×10⁻⁶ rad/mm"],
            ["Ultimate moment Mu", str(u.get("M","—"))+" kNm"],
            ["Ductility ratio μ",  str(mu)],
        ], [9*cm, 7.5*cm]))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=BORD))
    story.append(Paragraph(
        "SectionForge © 2026 Arash Nassirpour — Not a substitute for professional "
        "engineering judgement. All results must be verified by a qualified structural engineer.", S_note))

    doc.build(story)
    buf.seek(0)
    return buf


# ── EXCEL ─────────────────────────────────────────────────────────────
def generate_excel(params, curve, checks, mc_data):

    def hf():  return Font(bold=True, color="FFFFFF", size=11)
    def nf():  return Font(size=10)
    def bf(c="FFFFFF"): return PatternFill("solid", fgColor=c)
    def ca():  return Alignment(horizontal="center", vertical="center")
    def la():  return Alignment(horizontal="left",   vertical="center")
    def bdr():
        s = Side(style="thin", color="dde3ef")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws = wb.active;  ws.title = "Design Report"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    def sec_hdr(title, r):
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(r, 1, title)
        c.font = hf(); c.fill = bf("1a6cf5"); c.alignment = la()
        return r + 1

    def row(lbl, val, unit, r, shade=False, ok=None):
        for col, v in [(1,lbl),(2,val),(3,unit)]:
            c = ws.cell(r, col, v)
            c.font  = nf()
            c.fill  = bf("f0f3fa") if shade else bf()
            if col == 2 and ok is True:   c.fill = bf("d1fae5")
            if col == 2 and ok is False:  c.fill = bf("fee2e2")
            c.border    = bdr()
            c.alignment = ca() if col > 1 else la()
        return r + 1

    ws.merge_cells("A1:C1")
    ws.cell(1,1,"SectionForge — RC Design Report").font = Font(bold=True, size=16, color="1a6cf5")
    ws.cell(1,1).alignment = la()
    ws.merge_cells("A2:C2")
    ws.cell(2,1, f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}  ·  © 2026 Arash Nassirpour").font = Font(size=10, color="546080")
    r = 4

    r = sec_hdr("INPUT PARAMETERS", r)
    for i,(lbl,val,unit) in enumerate(_param_rows(params)):
        r = row(lbl, val, unit, r, i%2==1)

    r += 1
    uls = checks.get("uls",{}); sh  = checks.get("shear",{})
    sls = checks.get("sls",{}); rb  = checks.get("rebar",{})
    bal = curve.get("bal",{})
    r = sec_hdr("DESIGN CHECKS", r)
    sl = checks.get("slender") or {}
    st = sh.get("stirrups")
    extra = []
    if sl.get("enabled"):
        extra.append(("Slenderness λ / limit", f"{sl.get('lam','—')} / {sl.get('lam_lim','—')}", "—", None))
        if sl.get("magnified"):
            extra.append(("Magnified moment δM*", str(sl.get("Mc_kNm","—"))+" kNm", "—", None))
    if st:
        pick = st.get("pick")
        extra.append(("Shear links", f"{pick['legs']}-T{pick['dia']} @ {pick['s_mm']} mm" if pick else str(st.get("Asw_s_mm2_per_m","—"))+" mm²/m", "req.", None))
    for i,(lbl,val,unit,ok) in enumerate([
        ("ULS utilisation",   str(uls.get("util","—"))+"%",     "≤ 100%",    uls.get("pass",False)),
        ("φ factor",          str(uls.get("phi","—")),            "—",         None),
        ("NA depth",          str(uls.get("c_mm","—"))+" mm",    "—",         None),
        ("Shear φVc",         str(sh.get("phiVc_kN","—"))+" kN", "V*="+str(sh.get("Vd_kN","—")), sh.get("pass",False)),
        ("Crack width wk",    str(sls.get("wk_mm","—"))+" mm",   "≤ 0.3 mm",  sls.get("pass",False)),
        ("Steel area As",     str(rb.get("As_mm2","—"))+" mm²",  "min "+str(rb.get("As_min","—")), rb.get("pass_min",False)),
    ] + extra): r = row(lbl, val, unit, r, i%2==1, ok)

    # ── Sheet 2: M-N Curve ────────────────────────────────────────────
    ws2 = wb.create_sheet("M-N Curve Data")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22
    for col, hdr in [(1,"Moment φMn (kNm)"),(2,"Axial φPn (kN)")]:
        c = ws2.cell(1, col, hdr); c.font = hf(); c.fill = bf("1a6cf5"); c.alignment = ca()
    for i, pt in enumerate(curve.get("pts",[]), 2):
        fill = bf("f0f3fa") if i%2==0 else bf()
        for col, val in [(1,pt.get("x")),(2,pt.get("y"))]:
            c = ws2.cell(i,col,val); c.font=nf(); c.fill=fill; c.alignment=ca(); c.border=bdr()

    # ── Sheet 3: M-phi Curve ──────────────────────────────────────────
    if mc_data and mc_data.get("pts"):
        ws3 = wb.create_sheet("M-Phi Curve")
        ws3.column_dimensions["A"].width = 26
        ws3.column_dimensions["B"].width = 20
        for col, hdr in [(1,"Curvature φ (×10⁻⁶ rad/mm)"),(2,"Moment M (kNm)")]:
            c = ws3.cell(1, col, hdr); c.font = hf(); c.fill = bf("1a6cf5"); c.alignment = ca()
        for i, pt in enumerate(mc_data["pts"], 2):
            fill = bf("f0f3fa") if i%2==0 else bf()
            for col, val in [(1,pt.get("phi")),(2,pt.get("M"))]:
                c = ws3.cell(i,col,val); c.font=nf(); c.fill=fill; c.alignment=ca(); c.border=bdr()

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf
